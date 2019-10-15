# abrindo arquivo excel
from openpyxl import load_workbook

# a função load_workbook('arquivo.xlsx') retorna um objeto 
# book com os dados do arquivo excel. Para saber quais os nomes
# das planilhas existentes pode usar a propriedade book.sheetnames
book = load_workbook('sample.xlsx')
# print(book.sheetnames)

# obter o objeto sheet do arquivo aberto
sheet = book['Sheet']

#  recuperar valor de uma célula, basta se referenciar a célula
# e obter a propriedade "value" dela. 
valorA1 = sheet['A1'].value

print(valorA1)

# para saber quais as células ocupadas na planilha, é possivel
# obter os limites através da propriedade "dimensions" do objeto sheet.

# é possivel exibir o seu conteúdo, utilizando as dimensões
# junto com a estrutura de repetição "for" para poder ler todo
# o conteúdo da planilha de maneira dinâmica


d = sheet.dimensions

for c1, c2 in sheet[d]:
    print(c1.value+" "+str(c2.value))

    if type(c2.value) != str :
        soma = soma + c2.value

print("Soma total..: ", soma)
